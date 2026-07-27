"""Tests for the one-shot Excel export utility (not part of the exercise)."""

from __future__ import annotations

from pathlib import Path

import httpx
import pytest
import respx
from openpyxl import load_workbook

from src.models.domain.geometry import Point27700
from src.models.domain.gritbin import GritBin
from src.utils.excel_export import write_grit_bins_excel


def test_write_grit_bins_excel(tmp_path: Path) -> None:
    bins = [
        GritBin(
            title="GB-A",
            point=Point27700(1.0, 2.0),
            properties={
                "Subtitle": " Cinder Bank, Ironville, Alfreton ",
                "Street_Name": "Cinder Bank",
                "Town_Name": "Alfreton",
                "Area_Name": "Amber Valley Borough",
                "Location_Description": "On green area",
                "USRN": "600355",
            },
        ),
        GritBin(title="GB-B", point=Point27700(3.5, 4.5)),
    ]
    out = write_grit_bins_excel(bins, tmp_path / "gritbins.xlsx")
    assert out.exists()

    sheet = load_workbook(out).active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][:4] == ("title", "address", "postcode", "street_name")
    assert rows[1][0] == "GB-A"
    assert rows[1][1] == "Cinder Bank, Ironville, Alfreton"
    assert not rows[1][2]  # layer has no postcode attribute
    assert rows[1][3] == "Cinder Bank"
    assert rows[1][4] == "Alfreton"
    assert rows[2][0] == "GB-B"
    assert not rows[2][1]


@respx.mock
def test_export_excel_one_shot_guard(client, settings, tmp_path, monkeypatch) -> None:
    client.app.state.settings = settings
    target = tmp_path / "gritbins.xlsx"
    monkeypatch.setattr(
        "src.api.routers.gritbins.DEFAULT_EXPORT_PATH",
        target,
    )
    target.write_bytes(b"already-there")

    response = client.post("/grit-bins/export-excel")
    assert response.status_code == 409
    assert response.json()["error"]["code"] == "export_already_exists"


@respx.mock
def test_export_excel_writes_file(client, settings, tmp_path, monkeypatch) -> None:
    client.app.state.settings = settings
    target = tmp_path / "gritbins.xlsx"
    monkeypatch.setattr(
        "src.api.routers.gritbins.DEFAULT_EXPORT_PATH",
        target,
    )

    respx.get("https://wms.example.com/geoserver/DCC/ows").mock(
        return_value=httpx.Response(
            200,
            json={
                "type": "FeatureCollection",
                "features": [
                    {
                        "type": "Feature",
                        "geometry": {
                            "type": "Point",
                            "coordinates": [443620.0, 351800.0],
                        },
                        "properties": {
                            "Title": "GB-X",
                            "Subtitle": "Example Street, Derby",
                            "Street_Name": "Example Street",
                            "Town_Name": "Derby",
                        },
                    }
                ],
            },
        )
    )

    response = client.post("/grit-bins/export-excel")
    assert response.status_code == 200
    assert target.exists()
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    sheet = load_workbook(target).active
    row = list(sheet.iter_rows(values_only=True))[1]
    assert row[0] == "GB-X"
    assert row[1] == "Example Street, Derby"
    assert row[3] == "Example Street"
